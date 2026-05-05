from dataclasses import dataclass, field
from typing import List, Dict
import re
from openpyxl import Workbook


@dataclass
class Disciplina:
    turno: str
    codigo: str
    nome: str
    CH: int
    tipo: str
    horarios_aula: Dict[str, List[str]]
    periodo: int
    media: float
    requisito: List[str]
    co_requisito: List[str]
    

@dataclass
class Aluno:
    nome: str
    matricula: str
    periodos_integralizados: int
    disciplinas_concluidas: Dict[str, float]
    
@dataclass
class PERIODO:
    periodo_que_se_refere: str
    Segunda: dict = field(default_factory=lambda: {'M1': None, 'M2': None, 'M3': None, 'M4': None, 'M5': None, 'M6': None, 'T1': None, 'T2': None, 'T3': None, 'T4': None, 'T5': None, 'T6': None, 'N1': None, 'N2': None, 'N3': None})
    Terça: dict = field(default_factory=lambda: {'M1': None, 'M2': None, 'M3': None, 'M4': None, 'M5': None, 'M6': None, 'T1': None, 'T2': None, 'T3': None, 'T4': None, 'T5': None, 'T6': None, 'N1': None, 'N2': None, 'N3': None})
    Quarta: dict = field(default_factory=lambda: {'M1': None, 'M2': None, 'M3': None, 'M4': None, 'M5': None, 'M6': None, 'T1': None, 'T2': None, 'T3': None, 'T4': None, 'T5': None, 'T6': None, 'N1': None, 'N2': None, 'N3': None})
    Quinta: dict = field(default_factory=lambda: {'M1': Ncone, 'M2': None, 'M3': None, 'M4': None, 'M5': None, 'M6': None, 'T1': None, 'T2': None, 'T3': None, 'T4': None, 'T5': None, 'T6': None, 'N1': None, 'N2': None, 'N3': None})
    Sexta: dict = field(default_factory=lambda: {'M1': None, 'M2': None, 'M3': None, 'M4': None, 'M5': None, 'M6': None, 'T1': None, 'T2': None, 'T3': None, 'T4': None, 'T5': None, 'T6': None, 'N1': None, 'N2': None, 'N3': None})
    Sábado: dict = field(default_factory=lambda: {'M1': None, 'M2': None, 'M3': None, 'M4': None, 'M5': None, 'M6': None, 'T1': None, 'T2': None, 'T3': None, 'T4': None, 'T5': None, 'T6': None, 'N1': None, 'N2': None, 'N3': None})

banco_disciplinas = [
    Disciplina(turno = "T", codigo = "COMP359", nome = "Programação 1", CH = 72, tipo = "Obrigatória", periodo = 1 , horarios_aula = {"Sexta": ["T3", "T4", "T5", "T6"]}, media = 0, requisito = [], co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP360", nome = "Lógica para Programação", CH = 72, tipo = "Obrigatória", periodo = 1 , horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, requisito = [], co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP361", nome = "Computação, Sociedade e Ética", CH = 72, tipo = "Obrigatória", periodo = 1 , horarios_aula = {"Segunda": ["T5", "T6"], "Quarta": ["T5", "T6"]}, media = 0, requisito = [], co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP362", nome = "Matemática Discreta", CH = 72, tipo = "Obrigatória", periodo = 1 , horarios_aula = {"Terça": ["T1", "T2"], "Quinta": ["T1", "T2"]}, media = 0, requisito = [], co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP363", nome = "Cálculo Diferencial e Integral", CH = 144, tipo = "Obrigatória", periodo = 1 , horarios_aula = {"Terça": ["T3", "T4", "T5", "T6"], "Quinta": ["T3", "T4", "T5", "T6"]}, media = 0, requisito = [], co_requisito = []),

    Disciplina(turno = "M", codigo = "COMP364", nome = "Estrutura de Dados", CH = 72, tipo = "Obrigatória", periodo = 2, horarios_aula = {"Terça": ["M5", "M6"], "Quinta": ["M5", "M6"]}, media = 0, requisito = [], co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP365", nome = "Banco de Dados", CH = 72, tipo = "Obrigatória", periodo = 2, horarios_aula = {"Segunda": ["T1", "T2"], "Quarta": ["T1", "T2"]}, media = 0, requisito = [], co_requisito = []),
    Disciplina(turno = "M", codigo = "COMP366", nome = "Organização e Arquitetura de Computadores", CH = 72, tipo = "Obrigatória", periodo = 2, horarios_aula = {"Terça": ["M3", "M4"], "Quinta": ["M3", "M4"]}, media = 0, requisito = [], co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP367", nome = "Geometria Analítica", CH = 72, tipo = "Obrigatória", periodo = 2, horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, requisito = [], co_requisito = []),

    Disciplina(turno = "T", codigo = "COMP368", nome = "Redes de Computadores", CH = 72, tipo = "Obrigatória", periodo = 3, requisito = ["COMP359"], horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP369", nome = "Teoria dos Grafos", CH = 72, tipo = "Obrigatória", periodo = 3, requisito = ["COMP364", "COMP362"], horarios_aula = {"Segunda": ["T1", "T2"], "Quinta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "COMP370", nome = "Probabilidade e Estátistica", CH = 72, tipo = "Obrigatória", periodo = 3, requisito = ["COMP363"], horarios_aula = {"Segunda": ["M5", "M6"], "Quarta": ["M5", "M6"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "COMP371", nome = "Álgebra Linear", CH = 72, tipo = "Obrigatória", periodo = 3, requisito = ["COMP367"], horarios_aula = {"Segunda": ["M3", "M4"], "Quarta": ["M3", "M4"]}, media = 0, co_requisito = []),

    Disciplina(turno = "T", codigo = "COMP372", nome = "Programação 2", CH = 72, tipo = "Obrigatória", periodo = 4, requisito = ["COMP364", "COMP365", "COMP368"], horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = ["COMP373"]),
    Disciplina(turno = "T", codigo = "COMP373", nome = "Programação 3", CH = 72, tipo = "Obrigatória", periodo = 4, requisito = ["COMP364", "COMP365", "COMP368"], horarios_aula = {"Terça": ["T3", "T4"], "Quinta": ["T3", "T4"]}, media = 0, co_requisito = ["COMP372"]),
    Disciplina(turno = "T", codigo = "COMP374", nome = "Projeto e Análise de Algoritmos", CH = 72, tipo = "Obrigatória", periodo = 4, requisito = ["COMP364", "COMP369"], horarios_aula = {"Segunda": ["T5", "T6"], "Quinta": ["T5", "T6"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP376", nome = "Teoria da Computação", CH = 72, tipo = "Obrigatória", periodo = 4, requisito = [], horarios_aula = {"Segunda": ["T1", "T2"], "Quarta": ["T1", "T2"]}, media = 0, co_requisito = []),

    Disciplina(turno = "T", codigo = "COMP378", nome = "Sistemas Operacionais", CH = 72, tipo = "Obrigatória", periodo = 5, requisito = ["COMP366"], horarios_aula = {"Segunda": ["T1", "T2"], "Quarta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP379", nome = "Compiladores", CH = 72, tipo = "Obrigatória", periodo = 5, requisito = ["COMP364", "COMP376"], horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP380", nome = "Inteligência Artifical", CH = 72, tipo = "Obrigatória", periodo = 5, requisito = ["COMP360", "COMP364"], horarios_aula = {"Terça": ["T3", "T4"], "Quinta": ["T3", "T4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP381", nome = "Computação Gráfica", CH = 72, tipo = "Obrigatória", periodo = 5, requisito = [], horarios_aula = {"Terça": ["T5", "T6"], "Quinta": ["T5", "T6"]}, media = 0, co_requisito = []),

    Disciplina(turno = "M", codigo = "COMP382", nome = "Projeto e Desenvolvimento de Sistemas", CH = 288, tipo = "Obrigatória", periodo = 6, requisito = ["COMP359", "COMP360", "COMP361", "COMP362", "COMP363", "COMP364", "COMP365", "COMP366", "COMP367", "COMP368", "COMP369", "COMP370", "COMP371", "COMP372", "COMP373", "COMP374", "COMP376", "COMP378", "COMP379", "COMP380", "COMP381"], horarios_aula = {"Segunda": ["M3", "M4", "M5", "M6"], "Quarta": ["M3", "M4", "M5", "M6"], "Quinta": ["M5", "M6", "T1", "T2"]}, media = 0, co_requisito = []),

    Disciplina(turno = "T", codigo = "COMP386", nome = "Metodologia de Pesquisa e Trabalho Individual", CH = 72, tipo = "Obrigatória", periodo = 7, requisito = [], horarios_aula = {"Terça": ["T1", "T2"], "Quinta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP387", nome = "Noções de Direito", CH = 72, tipo = "Obrigatória", periodo = 7, requisito=[], horarios_aula = {"Terça": ["T3", "T4", "T5", "T6"]}, media = 0, co_requisito = []),



    Disciplina(turno = "T", codigo = "COMP377", nome = "ACE 1: Projeto I", CH = 75, tipo = "Obrigatória", periodo = 4, requisito = [], horarios_aula = {"Sexta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP383", nome = "ACE 2: Continuidade do Projeto", CH = 75, tipo = "Obrigatória", periodo = 4, requisito = ["COMP377"], horarios_aula = {"Sexta": ["T2", "T3"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP384", nome = "ACE 3: Projeto 2", CH = 73, tipo = "Obrigatória", periodo = 4, requisito = ["COMP377","COMP383"], horarios_aula = {"Quinta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "N", codigo = "COMP388", nome = "ACE 4: Continuidade do Projeto 2", CH = 75, tipo = "Obrigatória", periodo = 4, requisito = ["COMP377","COMP383","COMP384"], horarios_aula = {"Segunda": ["N1", "N2", "N3"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "COMP385", nome = "ACE 5: Evento", CH = 73, tipo = "Obrigatória", periodo = 4, requisito = ["COMP377","COMP383","COMP384","COMP388"], horarios_aula = {"Sábado": ["M2", "M3"]}, media = 0, co_requisito = []),



    Disciplina(turno = "T", codigo = "COMP404", nome = "Cálculo 3", CH = 72, tipo = "Ênfase", periodo = 8, requisito = ["COMP363"], horarios_aula = {"Segunda": ["T1", "T2"], "Quarta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP390", nome = "Aprendizagem de Máquina", CH = 72, tipo = "Ênfase", periodo = 8, requisito = ["COMP404"], horarios_aula = {"Segunda": ["T5", "T6"], "Quinta": ["T5", "T6"]}, media=0, co_requisito=[]),
    Disciplina(turno = "T", codigo = "COMP393", nome = "Redes Neurais e Aprendizado Profundo", CH = 72, periodo = 8, tipo = "Ênfase", requisito = [], horarios_aula = {"Terça": ["T1", "T2"], "Quinta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP396", nome = "Processamento Digital de Imagens", CH = 72, periodo = 8, tipo = "Ênfase", requisito = ["COMP381"], horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP400", nome = "Inteligência Artificial", CH = 72, tipo = "Ênfase", periodo = 8, requisito = [], horarios_aula = {"Terça": ["T3", "T4"], "Quinta": ["T3", "T4"]}, media = 0, co_requisito = []),

    Disciplina(turno = "T", codigo = "COMP397", nome = "Computação Evolucionária", CH = 72, tipo = "Ênfase", periodo = 8,requisito = [], horarios_aula = {"Terça": ["T3", "T4"], "Quinta": ["T3", "T4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP401", nome = "Ciência de Dados", CH = 72, tipo = "Ênfase", periodo = 8,requisito = ["COMP370"], horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = []),
    
    Disciplina(turno = "T", codigo = "COMP391", nome = "Sistemas Digitais", CH = 72, tipo = "Ênfase", periodo = 8, requisito = ["COMP404"], horarios_aula = {"Segunda": ["T1", "T2"], "Quarta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP398", nome = "Sistemas Embarcados", CH = 72, tipo = "Ênfase", periodo = 8, requisito = [], horarios_aula = {"Terça": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP402", nome = "Microcontroladores e Aplicações", CH = 72, tipo = "Ênfase", periodo = 8, requisito = [], horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = []),

    Disciplina(turno = "T", codigo = "COMP389", nome = "Conceitos de Linguagem de Programação", CH = 72, tipo = "Ênfase", periodo = 8, requisito = [], horarios_aula = {"Segunda": ["T3", "T4"], "Quarta": ["T3", "T4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP392", nome = "Sistemas Distribuídos", CH = 72, tipo = "Ênfase", periodo = 8, requisito = [], horarios_aula = {"Terça": ["T1", "T2"], "Quinta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "COMP399", nome = "Gerência de Projeto", CH = 72, tipo = "Ênfase", periodo = 8, requisito = ["COMP382"], horarios_aula = {"Terça": ["T3", "M4"], "Quinta": ["T3", "T4"]}, media = 0, co_requisito = []),  
    Disciplina(turno = "T", codigo = "COMP403", nome =  "Segurança de Sistemas Computacionais", CH = 72, tipo = "Ênfase", periodo = 8, requisito =["COMP368"], horarios_aula = {"Segunda": ["T1", "T2"], "Quarta": ["T1", "T2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "COMP395", nome = "Interação Homem-Máquina", CH = 72, tipo = "Ênfase", periodo = 8, requisito = ["COMP373"], horarios_aula = {"Segunda": ["M3", "M4"], "Quarta": ["M3", "M4"]}, media = 0, co_requisito = []),
    
    
    
    
    
    Disciplina(turno = "M", codigo = "CC1941", nome = "Cálculo 4", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Segunda": ["M1", "M2"], "Quarta": ["M1", "M2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "CC1945", nome = "Fundamentos de Libras", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Terça": ["M3", "M4"], "Quinta": ["M3", "M4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "CC1946", nome = "Geometria Computacional", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Segunda": ["M3", "M4"], "Quarta": ["M3", "M4"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "CC1958", nome = "Tópicos em Humanidades", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Terça": ["M1", "M2"], "Quinta": ["M1", "M2"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "COMP409", nome = "Tópicos em Matemática para Computação 1", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Segunda": ["M5", "M6"], "Quarta": ["M5", "M6"]}, media = 0, co_requisito = []),
    Disciplina(turno = "M", codigo = "COMP412", nome = "Tópicos em Física para Computação 1", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Terça": ["M5", "M6"], "Quinta": ["M5", "M6"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "CC1942", nome = "Cálculo Numérico", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Segunda": ["T5", "T6"], "Sexta": ["T5", "T6"]}, media = 0, co_requisito = []),
    Disciplina(turno = "T", codigo = "CC1943", nome = "Circuitos Digitais", CH = 72, tipo = "Eletiva", periodo = 0, requisito = [], horarios_aula = {"Terça": ["T5", "T6"], "Sexta": ["T5", "T6"]}, media = 0, co_requisito = []),
    
]

def menorNumDias(lista_disciplinas_a_fazer, ch_total, lista_de_periodos):
    dias_p_encaixar = []
    
    removerDisciplinas = []
    
    for discipli in banco_disciplinas:
        
        if discipli.codigo in lista_disciplinas_a_fazer and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and (not dias_p_encaixar or any(dia in dias_p_encaixar for dia in discipli.horarios_aula)):
            'insere agora essa disciplina'
            controle_dias = len(discipli.horarios_aula)
            pode_adicionar_disciplina = 0
            
            for dia, horario in discipli.horarios_aula.items():
                chaveDIA = dia
                controle_horarios = len(horario)
                pode_adicionar_horario = 0
                
                dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                
                
                for horarioDIA, codDisciplina in dicionario_do_dia.items():
                    
                    if horarioDIA in horario and codDisciplina is None:
                        pode_adicionar_horario += 1
                        
                        if pode_adicionar_horario == controle_horarios:
                            'significa que naquele dia específico os horários estão livre'
                            pode_adicionar_disciplina += 1
                            
                            if pode_adicionar_disciplina == controle_dias:
                                'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                'adiciona de fato a disciplina'
                                
                                if (ch_total + discipli.CH) > 540:
                                    for materia in removerDisciplinas:
                                        if materia in lista_disciplinas_a_fazer:
                                            lista_disciplinas_a_fazer.remove(materia)
                                    return
                                
                                else:
                                    for chave_dia in discipli.horarios_aula:
                                        if chave_dia not in dias_p_encaixar:
                                            dias_p_encaixar.append(chave_dia)
                                    
                                    ch_total += discipli.CH
                                    print("\nacessou a linha 163:\n", ch_total)
                                    print("materia adicionada:", discipli.codigo)
                                    removerDisciplinas.append(discipli.codigo)
                                    
                                    for d, h in discipli.horarios_aula.items():
                                        cDIA = d
                                        acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                        for hDIA, cod in acessDNV.items():
                                            for hDIA in h:
                                                    acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome

                                            '^^^ estou refazendo todo o processo já que não dária pra ficar adicionando já a disciplina quando pudesse porque um horário e/ou dia não permitisse mas ia ficar um fragmento daquela disciplina impedindo'

    for materia in removerDisciplinas:
        if materia in lista_disciplinas_a_fazer:
            lista_disciplinas_a_fazer.remove(materia)

def mesmoTurno(lista_disciplinas_a_fazer, ch_total, lista_de_periodos):
    turnO = None
   
    removerDisciplinas = []
    
    for discipli in banco_disciplinas:
        if discipli.codigo in lista_disciplinas_a_fazer and (discipli.tipo == 'Obrigatória' or discipli.tipo == 'Eletiva' or discipli.tipo == 'Ênfase') and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and (discipli.turno == turnO or turnO is None):
            'insere agora essa disciplina'
            controle_dias = len(discipli.horarios_aula)
            pode_adicionar_disciplina = 0
            
            for dia, horario in discipli.horarios_aula.items():
                chaveDIA = dia
                controle_horarios = len(horario)
                pode_adicionar_horario = 0
                
                dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                
                
                for horarioDIA, codDisciplina in dicionario_do_dia.items():
                    
                    if horarioDIA in horario and codDisciplina is None:
                        pode_adicionar_horario += 1
                        
                        if pode_adicionar_horario == controle_horarios:
                            'significa que naquele dia específico os horários estão livre'
                            pode_adicionar_disciplina += 1
                            
                            if pode_adicionar_disciplina == controle_dias:
                                'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                'adiciona de fato a disciplina'
                                
                                if (ch_total + discipli.CH) > 540:
                                    for materia in removerDisciplinas:
                                        if materia in lista_disciplinas_a_fazer:
                                            lista_disciplinas_a_fazer.remove(materia)
                                    return
                                
                                else:
                                    if turnO is None:
                                        turnO = discipli.turno
                                        
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 163:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome
                                    else:
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 163:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome

                                            '^^^ estou refazendo todo o processo já que não dária pra ficar adicionando já a disciplina quando pudesse porque um horário e/ou dia não permitisse mas ia ficar um fragmento daquela disciplina impedindo'

    for materia in removerDisciplinas:
        if materia in lista_disciplinas_a_fazer:
            lista_disciplinas_a_fazer.remove(materia)
            
def auxilia_max3_td_dia(lista_disciplinas_a_fazer, ch_total, lista_de_periodos, DIAAA):
    print("\nACESSOU 247 INICIO AUXILIA_MAX3, CH_TOTAL ===\n", ch_total)
    
    
    tem_que_sair = 0
    removerDisciplinas = []
    
    if DIAAA == 2:
        #segunda
        for discipli in banco_disciplinas:
            if tem_que_sair == 1:
                break
            
            if discipli.codigo in lista_disciplinas_a_fazer and (discipli.tipo == 'Obrigatória' or discipli.tipo == 'Eletiva' or discipli.tipo == 'Ênfase') and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and any(chave == "Segunda" for chave in discipli.horarios_aula):
                'insere agora essa disciplina'
                controle_dias = len(discipli.horarios_aula)
                pode_adicionar_disciplina = 0
                
                for dia, horario in discipli.horarios_aula.items():
                    if tem_que_sair == 1:
                        break
                    chaveDIA = dia
                    controle_horarios = len(horario)
                    pode_adicionar_horario = 0
                    
                    dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                    
                    
                    for horarioDIA, codDisciplina in dicionario_do_dia.items():
                        if tem_que_sair == 1:
                            break
                        
                        if horarioDIA in horario and codDisciplina is None:
                            pode_adicionar_horario += 1
                            
                            if pode_adicionar_horario == controle_horarios:
                                'significa que naquele dia específico os horários estão livre'
                                pode_adicionar_disciplina += 1
                                
                                if pode_adicionar_disciplina == controle_dias:
                                    'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                    'adiciona de fato a disciplina'
                                    
                                    if (ch_total + discipli.CH) > 540:
                                        print("\nACESSOU 290\n", ch_total)
                                        for materia in removerDisciplinas:
                                            if materia in lista_disciplinas_a_fazer:
                                                lista_disciplinas_a_fazer.remove(materia)
                                        return "SAIA"
                                    
                                    else:
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 298 SEGUNDA CH_TOTAL:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome
                                            
                                        '''aqui precisa ter algo que saia de todo esse for aninhado, se vai ser break ou return, n sei mas acho que é break'''
                                        tem_que_sair = 1
        
        for materia in removerDisciplinas:
            if materia in lista_disciplinas_a_fazer:
                lista_disciplinas_a_fazer.remove(materia)
                
        return ch_total
        
    elif DIAAA == 3:
        #terça
        for discipli in banco_disciplinas:
            if tem_que_sair == 1:
                break
            
            if discipli.codigo in lista_disciplinas_a_fazer and (discipli.tipo == 'Obrigatória' or discipli.tipo == 'Eletiva' or discipli.tipo == 'Ênfase') and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and any(chave == "Terça" for chave in discipli.horarios_aula):
                'insere agora essa disciplina'
                controle_dias = len(discipli.horarios_aula)
                pode_adicionar_disciplina = 0
                
                for dia, horario in discipli.horarios_aula.items():
                    if tem_que_sair == 1:
                        break
                    chaveDIA = dia
                    controle_horarios = len(horario)
                    pode_adicionar_horario = 0
                    
                    dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                    
                    
                    for horarioDIA, codDisciplina in dicionario_do_dia.items():
                        if tem_que_sair == 1:
                            break
                        
                        if horarioDIA in horario and codDisciplina is None:
                            pode_adicionar_horario += 1
                            
                            if pode_adicionar_horario == controle_horarios:
                                'significa que naquele dia específico os horários estão livre'
                                pode_adicionar_disciplina += 1
                                
                                if pode_adicionar_disciplina == controle_dias:
                                    'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                    'adiciona de fato a disciplina'
                                    
                                    if (ch_total + discipli.CH) > 540:
                                        for materia in removerDisciplinas:
                                            if materia in lista_disciplinas_a_fazer:
                                                lista_disciplinas_a_fazer.remove(materia)
                                        return "SAIA"
                                    
                                    else:
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 362 TERÇA CH_TOTAL:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome
                                            
                                        '''aqui precisa ter algo que saia de todo esse for aninhado, se vai ser break ou return, n sei mas acho que é break'''
                                        tem_que_sair = 1
        
        for materia in removerDisciplinas:
            if materia in lista_disciplinas_a_fazer:
                lista_disciplinas_a_fazer.remove(materia)
                
        return ch_total
        
        
    elif DIAAA == 4:
        #quarta
        for discipli in banco_disciplinas:
            if tem_que_sair == 1:
                break
            
            if discipli.codigo in lista_disciplinas_a_fazer and (discipli.tipo == 'Obrigatória' or discipli.tipo == 'Eletiva' or discipli.tipo == 'Ênfase') and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and any(chave == "Quarta" for chave in discipli.horarios_aula):
                'insere agora essa disciplina'
                controle_dias = len(discipli.horarios_aula)
                pode_adicionar_disciplina = 0
                
                for dia, horario in discipli.horarios_aula.items():
                    if tem_que_sair == 1:
                        break
                    chaveDIA = dia
                    controle_horarios = len(horario)
                    pode_adicionar_horario = 0
                    
                    dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                    
                    
                    for horarioDIA, codDisciplina in dicionario_do_dia.items():
                        if tem_que_sair == 1:
                            break
                        
                        if horarioDIA in horario and codDisciplina is None:
                            pode_adicionar_horario += 1
                            
                            if pode_adicionar_horario == controle_horarios:
                                'significa que naquele dia específico os horários estão livre'
                                pode_adicionar_disciplina += 1
                                
                                if pode_adicionar_disciplina == controle_dias:
                                    'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                    'adiciona de fato a disciplina'
                                    
                                    if (ch_total + discipli.CH) > 540:
                                        for materia in removerDisciplinas:
                                            if materia in lista_disciplinas_a_fazer:
                                                lista_disciplinas_a_fazer.remove(materia)
                                        return "SAIA"
                                    
                                    else:
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 427, QUARTA CHTOTAL:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome
                                            
                                        '''aqui precisa ter algo que saia de todo esse for aninhado, se vai ser break ou return, n sei mas acho que é break'''
                                        tem_que_sair = 1
        
        for materia in removerDisciplinas:
            if materia in lista_disciplinas_a_fazer:
                lista_disciplinas_a_fazer.remove(materia)
                
        return ch_total
        
        
    elif DIAAA == 5:
        #quinta
        for discipli in banco_disciplinas:
            if tem_que_sair == 1:
                break
            
            if discipli.codigo in lista_disciplinas_a_fazer and (discipli.tipo == 'Obrigatória' or discipli.tipo == 'Eletiva' or discipli.tipo == 'Ênfase') and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and any(chave == "Quinta" for chave in discipli.horarios_aula):
                'insere agora essa disciplina'
                controle_dias = len(discipli.horarios_aula)
                pode_adicionar_disciplina = 0
                
                for dia, horario in discipli.horarios_aula.items():
                    if tem_que_sair == 1:
                        break
                    chaveDIA = dia
                    controle_horarios = len(horario)
                    pode_adicionar_horario = 0
                    
                    dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                    
                    
                    for horarioDIA, codDisciplina in dicionario_do_dia.items():
                        if tem_que_sair == 1:
                            break
                        
                        if horarioDIA in horario and codDisciplina is None:
                            pode_adicionar_horario += 1
                            
                            if pode_adicionar_horario == controle_horarios:
                                'significa que naquele dia específico os horários estão livre'
                                pode_adicionar_disciplina += 1
                                
                                if pode_adicionar_disciplina == controle_dias:
                                    'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                    'adiciona de fato a disciplina'
                                    
                                    if (ch_total + discipli.CH) > 540:
                                        for materia in removerDisciplinas:
                                            if materia in lista_disciplinas_a_fazer:
                                                lista_disciplinas_a_fazer.remove(materia)
                                        return "SAIA"
                                    
                                    else:
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 492, QUINTA CHTOTAL:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome
                                            
                                        '''aqui precisa ter algo que saia de todo esse for aninhado, se vai ser break ou return, n sei mas acho que é break'''
                                        tem_que_sair = 1
        
        for materia in removerDisciplinas:
            if materia in lista_disciplinas_a_fazer:
                lista_disciplinas_a_fazer.remove(materia)
                
        return ch_total
        
        
    elif DIAAA == 6:
        #sexta
        for discipli in banco_disciplinas:
            if tem_que_sair == 1:
                break
            
            if discipli.codigo in lista_disciplinas_a_fazer and (discipli.tipo == 'Obrigatória' or discipli.tipo == 'Eletiva' or discipli.tipo == 'Ênfase') and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and any(chave == "Sexta" for chave in discipli.horarios_aula):
                'insere agora essa disciplina'
                controle_dias = len(discipli.horarios_aula)
                pode_adicionar_disciplina = 0
                
                for dia, horario in discipli.horarios_aula.items():
                    if tem_que_sair == 1:
                        break
                    chaveDIA = dia
                    controle_horarios = len(horario)
                    pode_adicionar_horario = 0
                    
                    dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                    
                    
                    for horarioDIA, codDisciplina in dicionario_do_dia.items():
                        if tem_que_sair == 1:
                            break
                        
                        if horarioDIA in horario and codDisciplina is None:
                            pode_adicionar_horario += 1
                            
                            if pode_adicionar_horario == controle_horarios:
                                'significa que naquele dia específico os horários estão livre'
                                pode_adicionar_disciplina += 1
                                
                                if pode_adicionar_disciplina == controle_dias:
                                    'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                    'adiciona de fato a disciplina'
                                    
                                    if (ch_total + discipli.CH) > 540:
                                        for materia in removerDisciplinas:
                                            if materia in lista_disciplinas_a_fazer:
                                                lista_disciplinas_a_fazer.remove(materia)
                                        return "SAIA"
                                    
                                    else:
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 557 SEXTA, CHTOTAL:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome
                                            
                                        '''aqui precisa ter algo que saia de todo esse for aninhado, se vai ser break ou return, n sei mas acho que é break'''
                                        tem_que_sair = 1
        
        for materia in removerDisciplinas:
            if materia in lista_disciplinas_a_fazer:
                lista_disciplinas_a_fazer.remove(materia)
                
        return ch_total
        
        
    elif DIAAA == 7:
        #sabado
        for discipli in banco_disciplinas:
            if tem_que_sair == 1:
                break
            
            if discipli.codigo in lista_disciplinas_a_fazer and (discipli.tipo == 'Obrigatória' or discipli.tipo == 'Eletiva' or discipli.tipo == 'Ênfase') and all(item not in lista_disciplinas_a_fazer for item in discipli.requisito) and any(chave == "Sábado" for chave in discipli.horarios_aula):
                'insere agora essa disciplina'
                controle_dias = len(discipli.horarios_aula)
                pode_adicionar_disciplina = 0
                
                for dia, horario in discipli.horarios_aula.items():
                    if tem_que_sair == 1:
                        break
                    chaveDIA = dia
                    controle_horarios = len(horario)
                    pode_adicionar_horario = 0
                    
                    dicionario_do_dia = getattr(lista_de_periodos[-1], chaveDIA)
                    
                    
                    for horarioDIA, codDisciplina in dicionario_do_dia.items():
                        if tem_que_sair == 1:
                            break
                        
                        if horarioDIA in horario and codDisciplina is None:
                            pode_adicionar_horario += 1
                            
                            if pode_adicionar_horario == controle_horarios:
                                'significa que naquele dia específico os horários estão livre'
                                pode_adicionar_disciplina += 1
                                
                                if pode_adicionar_disciplina == controle_dias:
                                    'significa que em todos os dias, e seus respectivos horarios, a disciplina ta livre'
                                    'adiciona de fato a disciplina'
                                    
                                    if (ch_total + discipli.CH) > 540:
                                        for materia in removerDisciplinas:
                                            if materia in lista_disciplinas_a_fazer:
                                                lista_disciplinas_a_fazer.remove(materia)
                                        return "SAIA"
                                    
                                    else:
                                        ch_total += discipli.CH
                                        print("\nacessou a linha 622 - SÁBADO, CH TOTAL:\n", ch_total)
                                        print("materia adicionada:", discipli.codigo)
                                        removerDisciplinas.append(discipli.codigo)
                                        
                                        for d, h in discipli.horarios_aula.items():
                                            cDIA = d
                                            acessDNV = getattr(lista_de_periodos[-1], cDIA)
                                            for hDIA, cod in acessDNV.items():
                                                for hDIA in h:
                                                        acessDNV[hDIA] = discipli.codigo + "-" + discipli.nome
                                            
                                        '''aqui precisa ter algo que saia de todo esse for aninhado, se vai ser break ou return, n sei mas acho que é break'''
                                        tem_que_sair = 1
        
        for materia in removerDisciplinas:
            if materia in lista_disciplinas_a_fazer:
                lista_disciplinas_a_fazer.remove(materia)
                
        return ch_total
        
   
def max3_td_dia(lista_disciplinas_a_fazer, ch_total, lista_de_periodos):
    tem_que_sair = 0
    removerDisciplinas = []
    ch_total_atualizado = 0
    auxcontro = 0
    print("\nACESSOU 648 - inicio max3_td_dia, valor de ch_total\n", ch_total)
    
    while auxcontro < 3:
        print("\nACESSOU 651 - segunda\n", ch_total_atualizado)
        ch_total_atualizado = auxilia_max3_td_dia(lista_disciplinas_a_fazer, ch_total_atualizado, lista_de_periodos, 2)
        if ch_total_atualizado == "SAIA":
            return
        
        else:
            print("\nACESSOU 657 - terça\n", ch_total_atualizado)
            ch_total_atualizado = auxilia_max3_td_dia(lista_disciplinas_a_fazer, ch_total_atualizado, lista_de_periodos, 3)
            if ch_total_atualizado == "SAIA":
                return
            
            else:
                print("\nACESSOU 663 - quarta\n", ch_total_atualizado)
                ch_total_atualizado = auxilia_max3_td_dia(lista_disciplinas_a_fazer, ch_total_atualizado, lista_de_periodos, 4)
                if ch_total_atualizado == "SAIA":
                    return
                
                else:
                    print("\nACESSOU 669 - quinta\n", ch_total_atualizado)
                    ch_total_atualizado = auxilia_max3_td_dia(lista_disciplinas_a_fazer, ch_total_atualizado, lista_de_periodos, 5)
                    if ch_total_atualizado == "SAIA":
                        return
                    
                    else:
                        print("\nACESSOU 675 - sexta\n", ch_total_atualizado)
                        ch_total_atualizado = auxilia_max3_td_dia(lista_disciplinas_a_fazer, ch_total_atualizado, lista_de_periodos, 6)
                        if ch_total_atualizado == "SAIA":
                            return
                        
                        else:
                            print("\nACESSOU 681 - sábado\n", ch_total_atualizado)
                            ch_total_atualizado = auxilia_max3_td_dia(lista_disciplinas_a_fazer, ch_total_atualizado, lista_de_periodos, 7)
                            if ch_total_atualizado == "SAIA":
                                return
                        
        
        print("\nACESSOU 687 dE UMA ITERAÇÃO WHILE:\n", ch_total)
        auxcontro += 1  
    

def inserePeriodo(escolhas_dicio, escolha_enfase, lista_disciplinas_a_fazer, periodos_feitos, lista_enfase, lista_alunos, list_tipoEnfase_feito):
    periodo_max = 0
    ennnfase = "Nenhuma"
    
    if escolhas_dicio[1] == 0:
        periodo_max = 12
        
    elif escolhas_dicio == 1:
        periodo_max = 10
    else:
        periodo_max = 8
        
    total_periodo_escolhido = 0
    
    if escolhas_dicio[2] == 0:
        total_periodo_escolhido = 8
    elif escolhas_dicio[2] == 1:
        total_periodo_escolhido = periodo_max
    else:
        total_periodo_escolhido = (periodo_max + 8) / 2
    
    aux = 1
    
    if escolhas_dicio[3] == 0:
        print("\nEscolha a ênfase desejada:\n")
        
        for escolha in escolha_enfase.keys():
            print(aux, "-",escolha,"\n")
            aux += 1
        
        while True:
            try:
                escolha = int(input("Digite o número da ênfase desejada: "))
                if escolha > 4 or escolha < 1:
                    print("ERRO: DIGITE 1, 2, 3 OU 4!\n")
                else:
                    if escolha == 1:
                        escolha = "COMPUTAÇÃO VISUAL"
                        ennnfase = escolha
                        
                    elif escolha == 2:
                        escolha = "SISTEMAS INTELIGENTES"
                        ennnfase = escolha

                    elif escolha == 3:
                        escolha = "SISTEMAS DE COMPUTAÇÃO"
                        ennnfase = escolha

                    else:
                        escolha = "SISTEMAS DE INFORMAÇÃO"
                        ennnfase = escolha
                    
                    break
            except:
                print("ERRO: INSIRA SOMENTE NÚMEROS INTEIROS!\n")
        
        for chave, valor in escolha_enfase.items():
            if escolha == chave:
                for valores in valor:
                    if valores not in lista_disciplinas_a_fazer and valores not in list_tipoEnfase_feito:
                        lista_disciplinas_a_fazer.append(valores)
                
                
                
        print("linha 229 enfase escolhida:", lista_disciplinas_a_fazer)

    elif escolhas_dicio[3] == 1:
        en1 = 5
        en2 = 5
        en3 = 5
        en4 = 5
        
        list_de_list = [escolha_enfase["COMPUTAÇÃO VISUAL"], escolha_enfase["SISTEMAS INTELIGENTES"], escolha_enfase["SISTEMAS DE COMPUTAÇÃO"], escolha_enfase["SISTEMAS DE INFORMAÇÃO"]]
        
        for sublista in list_de_list:
            for item in sublista:
                if item in lista_enfase and sublista == escolha_enfase["COMPUTAÇÃO VISUAL"]:
                    en1 -= 1
                    
                if item in lista_enfase and sublista == escolha_enfase["SISTEMAS INTELIGENTES"]:
                    en2 -= 1
                    
                if item in lista_enfase and sublista == escolha_enfase["SISTEMAS DE COMPUTAÇÃO"]:
                    en3 -= 1
                    
                if item in lista_enfase and sublista == escolha_enfase["SISTEMAS DE INFORMAÇÃO"]:
                    en4 -= 1
        
        auxauxaux = 0
        
        escol = 6
        
        if escol > en1:
            escol = en1
            auxauxaux = 1
            
        if escol > en2:
            escol = en2
            auxauxaux = 2
            
        if escol > en3:
            escol = en3
            auxauxaux = 3
        if escol > en4:
            escol = en4
            auxauxaux = 4
        
        if auxauxaux == 1:
            ennnfase = "COMPUTAÇÃO VISUAL"

            for valor in escolha_enfase["COMPUTAÇÃO VISUAL"]:
                if valor not in lista_enfase and valor not in list_tipoEnfase_feito and valor not in lista_disciplinas_a_fazer:
                    lista_disciplinas_a_fazer.append(valor)
                    
        elif auxauxaux == 2:
            ennnfase = "SISTEMAS INTELIGENTES"

            for valor in escolha_enfase["SISTEMAS INTELIGENTES"]:
                if valor not in lista_enfase and valor not in list_tipoEnfase_feito and valor not in lista_disciplinas_a_fazer:
                    lista_disciplinas_a_fazer.append(valor)
                    
        elif auxauxaux == 3:
            ennnfase = "SISTEMAS COMPUTAÇÃO"

            for valor in escolha_enfase["SISTEMAS COMPUTAÇÃO"]:
                if valor not in lista_enfase and valor not in list_tipoEnfase_feito and valor not in lista_disciplinas_a_fazer:
                    lista_disciplinas_a_fazer.append(valor)
                    
        elif auxauxaux == 4:
            ennnfase = "SISTEMAS INFORMAÇÃO"

            for valor in escolha_enfase["SISTEMAS INFORMAÇÃO"]:
                if valor not in lista_enfase and valor not in list_tipoEnfase_feito and valor not in lista_disciplinas_a_fazer:
                    lista_disciplinas_a_fazer.append(valor)
        
        print("\n\n LINHA 420 \n\n", lista_disciplinas_a_fazer)
    
    
    periodos_feitos += 1
    lista_de_periodos = []
    ch_total = 0
    
    condicao = ""
    
    if escolhas_dicio[4] == 0:
        condicao = "menor_n_dias"
        
        while periodos_feitos <= total_periodo_escolhido:
            lista_de_periodos.append(PERIODO(periodo_que_se_refere = periodos_feitos))
        
            
            menorNumDias(lista_disciplinas_a_fazer, ch_total, lista_de_periodos)
            periodos_feitos += 1
            
            print("acessou a linha 1495:")
            print("\n",lista_de_periodos[-1],"\n")
        
        if total_periodo_escolhido != 12:
            while periodos_feitos <= 12:
                lista_de_periodos.append(PERIODO(periodo_que_se_refere = periodos_feitos))
        
            
                menorNumDias(lista_disciplinas_a_fazer, ch_total, lista_de_periodos)
                lista_de_periodos[-1].periodo_que_se_refere = f"{periodos_feitos} - EXCEDENTE"
                periodos_feitos += 1
                

                print("acessou a linha 1495:")
                print("\n",lista_de_periodos[-1],"\n")
        
        
    elif escolhas_dicio[4] == 1:
        condicao = "mesmo_turno"
        
        while periodos_feitos <= total_periodo_escolhido:
            lista_de_periodos.append(PERIODO(periodo_que_se_refere = periodos_feitos))
        
            
            mesmoTurno(lista_disciplinas_a_fazer, ch_total, lista_de_periodos)
            periodos_feitos += 1
            
            print("acessou a linha 1495:")
            print("\n",lista_de_periodos[-1],"\n")

        if total_periodo_escolhido != 12:
            while periodos_feitos <= 12:
                lista_de_periodos.append(PERIODO(periodo_que_se_refere = periodos_feitos))
        
            
                mesmoTurno(lista_disciplinas_a_fazer, ch_total, lista_de_periodos)
                lista_de_periodos[-1].periodo_que_se_refere = f"{periodos_feitos} - EXCEDENTE"
                periodos_feitos += 1
                

                print("acessou a linha 1495:")
                print("\n",lista_de_periodos[-1],"\n")
        
    else:
        condicao = "max3d-seg-sex"
        
        while periodos_feitos <= total_periodo_escolhido:
            lista_de_periodos.append(PERIODO(periodo_que_se_refere = periodos_feitos))
            
            max3_td_dia(lista_disciplinas_a_fazer, ch_total, lista_de_periodos)
            periodos_feitos += 1
            
            print("acessou a linha 1495:")
            print("\n",lista_de_periodos[-1],"\n")

        if total_periodo_escolhido != 12:
            while periodos_feitos <= 12:
                lista_de_periodos.append(PERIODO(periodo_que_se_refere = periodos_feitos))
        
            
                max3_td_dia(lista_disciplinas_a_fazer, ch_total, lista_de_periodos)
                lista_de_periodos[-1].periodo_que_se_refere = f"{periodos_feitos} - EXCEDENTE"
                periodos_feitos += 1
                

                print("acessou a linha 1495:")
                print("\n",lista_de_periodos[-1],"\n")
        
    
    
    for objeto in lista_de_periodos:
        print('\n', objeto, '\n')
    
    mapeamento_colunas = {
        "M1": "B",
        "M2": "C",
        "M3": "D",
        "M4": "E",
        "M5": "F",
        "M6": "G",
        "T1": "H",
        "T2": "I",
        "T3": "J",
        "T4": "K",
        "T5": "L",
        "T6": "M",
        "N1": "N",
        "N2": "O",
        "N3": "P"
    }
    mapeamento_numero_letra = {
        2: "B",
        3: "C",
        4: "D",
        5: "E",
        6: "F",
        7: "G",
        8: "H",
        9: "I",
        10: "J",
        11: "K",
        12: "L",
        13: "M",
        14: "N",
        15: "O",
        16: "P",
        17: "R",
        18: "S",
        19: "T",
        20: "U",
        21: "V",
        22: "W",
        23: "Y",
        24: "X",
        25: "Z"
    }



    wb = Workbook()
    ws = wb.active

    ws["A1"] = "NOME"
    ws["B1"] = lista_alunos[0].nome
    ws["D1"] = "MAX periodo"
    ws["E1"] = periodo_max
    ws["G1"] = "ÊNFASE"
    ws["H1"] = ennnfase
    ws["A2"] = "MATRICULA"
    ws["B2"] = lista_alunos[0].matricula
    ws["D2"] = "MAX escolhido"
    ws["E2"] = total_periodo_escolhido
    ws["G2"] = "CONDIÇÃO"
    ws["H2"] = condicao

    linha = 4

    for qt in lista_de_periodos:
        ws[f"A{linha}"] = "PERIODO"
        ws[f"B{linha}"] = qt.periodo_que_se_refere

        linha += 2
        ws[f"B{linha}"] = "M1"
        ws[f"C{linha}"] = "M2"
        ws[f"D{linha}"] = "M3"
        ws[f"E{linha}"] = "M4"
        ws[f"F{linha}"] = "M5"
        ws[f"G{linha}"] = "M6"
        ws[f"H{linha}"] = "T1"
        ws[f"I{linha}"] = "T2"
        ws[f"J{linha}"] = "T3"
        ws[f"K{linha}"] = "T4"
        ws[f"L{linha}"] = "T5"
        ws[f"M{linha}"] = "T6"
        ws[f"N{linha}"] = "N1"
        ws[f"O{linha}"] = "N2"
        ws[f"P{linha}"] = "N3"

        linha += 1
        ws[f"A{linha}"] = "SEGUNDA"

        for chave, valor in qt.Segunda.items():
            if valor is not None and chave in mapeamento_colunas:
                coluna = mapeamento_colunas[chave]
                ws[f"{coluna}{linha}"] = valor

        linha += 1
        ws[f"A{linha}"] = "TERÇA"

        for chave, valor in qt.Terça.items():
            if valor is not None and chave in mapeamento_colunas:
                coluna = mapeamento_colunas[chave]
                ws[f"{coluna}{linha}"] = valor


        linha += 1
        ws[f"A{linha}"] = "QUARTA"

        for chave, valor in qt.Quarta.items():
            if valor is not None and chave in mapeamento_colunas:
                coluna = mapeamento_colunas[chave]
                ws[f"{coluna}{linha}"] = valor


        linha += 1
        ws[f"A{linha}"] = "QUINTA"

        for chave, valor in qt.Quinta.items():
            if valor is not None and chave in mapeamento_colunas:
                coluna = mapeamento_colunas[chave]
                ws[f"{coluna}{linha}"] = valor


        linha += 1
        ws[f"A{linha}"] = "SEXTA"

        for chave, valor in qt.Sexta.items():
            if valor is not None and chave in mapeamento_colunas:
                coluna = mapeamento_colunas[chave]
                ws[f"{coluna}{linha}"] = valor

        linha += 1
        ws[f"A{linha}"] = "SÁBADO"

        for chave, valor in qt.Sábado.items():
            if valor is not None and chave in mapeamento_colunas:
                coluna = mapeamento_colunas[chave]
                ws[f"{coluna}{linha}"] = valor
        
        linha += 2

    ws[f"A{linha}"] = "RESTOU DISCIPLINA?"
    contauxilia = 1

    for qt in lista_disciplinas_a_fazer:
        contauxilia += 1
        for objeto in banco_disciplinas:
            if qt == objeto.codigo:
                for chave, valor in mapeamento_numero_letra.items():
                    if contauxilia == chave:
                        ws[f"{valor}{linha}"] = qt + " - " + objeto.nome
                        break
                break


    wb.save("planilha.xlsx")

    



def calculaNome2(parte, escolhas_dicio, etapa):
    letra_numero = {"Q" : 1, "A" : 2, "Z" : 3, "W" : 6, "S" : 5, "X" : 4, "E" : 7, "D" : 8, "C" : 9, "R" : 6, "F" : 7, "V" : 8, "T" : 5, "G" : 4, "B" : 3, "Y" : 2, "H" : 1, "N" : 2, "U" : 3, "J" : 4, "M" : 5, "I" : 8, "K" : 7, "Ó" : 6, "O" : 9, "L" : 8, "P" : 4, "Ç" : 5, "Ô" : 6, "Á" : 3, "É" : 2, "Â" : 1, "Ã" : 4, "Í" : 3, "Ê" : 2}
    soma = 0
    
    for letra in parte:
        if letra in letra_numero:
            soma += letra_numero[letra]
            
    
    if soma % 3 <= 0:
        escolhas_dicio[etapa] = 0
    elif soma % 3 == 1:
        escolhas_dicio[etapa] = 1
    elif soma % 3 >= 2:
        escolhas_dicio[etapa] = 2
   


def calculaNome1(nome: str, escolhas_dicio):


    'print(nome)'
    partes = nome.split()
    partes_sem_lig = []
    
    for parte in partes:
        if parte not in ["DE", "DO", "DA", "DOS", "DAS", "DES"]:
            partes_sem_lig.append(parte)
            
    
    print(partes_sem_lig)
    
    
    tamanho = len(partes_sem_lig)
    'print(tamanho)'
    

    aux = 1


    while(aux < 5):
        if tamanho >= 4:
            calculaNome2(partes_sem_lig[aux-1], escolhas_dicio, aux)
            aux += 1
            
        elif tamanho == 3:
            if aux == 4:
                calculaNome2(partes_sem_lig[0], escolhas_dicio, aux)
                aux += 1
            else:
                calculaNome2(partes_sem_lig[aux-1], escolhas_dicio, aux)
                aux += 1
                
        elif tamanho == 2:
            if aux == 4:
                calculaNome2(partes_sem_lig[1], escolhas_dicio, aux)
                aux += 1
            elif aux == 3:
                calculaNome2(partes_sem_lig[0], escolhas_dicio, aux)
                aux += 1
            else:
                calculaNome2(partes_sem_lig[aux-1], escolhas_dicio, aux)
                aux += 1
                
        elif tamanho == 1:
            calculaNome2(partes_sem_lig[0], escolhas_dicio, aux)
            aux += 1


def main():
    lista_alunos = [] #receberá objetos da classe aluno instanciados (criados) aqui nessa lista
    QT_alunos = 0 #quantidade de alunos a serem aconselhados (primeiramente criados para depois serem aconselhados)

    #isso daqui é tipo o while(1), onde só vai sair dessa condição quando tiver um break (quando der certo no bloco try except)'''
    while True: 
        try:
            QT_alunos = int(input("Insira a quantidade de alunos a serem aconselhados\n(números inteiros positivos): "))
            if QT_alunos > 0:
                break
            else:
                print("\n!ERRO!, insira SOMENTE números inteiros positivos!\n")
        except:
            print("\n!ERRO!, insira SOMENTE números inteiros positivos!\n")

    QT_alunos_aux_controle = 0
    aux_nome_codigo = "" #vai entrar no dicionário aux_codigo_media_disciplina como chave
    aux_nome = ""
    aux_matricula = ""
    aux_codigo_media_disciplina = {}
    aux_media = 0.0 #vai entrar no dicionário aux_codigo_media_disciplina como valor
    aux_periodos = 0

    #inserção de informações de alunos
    while QT_alunos_aux_controle < QT_alunos:
            print("\n------Inserção de informações de Aluno:", QT_alunos_aux_controle+1,"------\n")
            
            #insere o nome
            while True:
                try:
                    aux_nome = input("Insira nome do aluno\n(somente letras maíusculas): ")

                    if re.match(r'^[A-ZÀ-Ý\s]+$', aux_nome):  # Aceita letras, espaços e hífens <---- estudar melhor isso e a biblioteca re
                        break

                    else:
                        print("\n!ERRO!: insira SOMENTE LETRAS MAIÚSCULAS!\n")

                except:
                    print("\n!ERRO!, insira SOMENTE LETRAS MAIÚSCULAS!\n")

            #insere a matricula
            while True:
                try:
                    aux_matricula = input("\nInsira a matricula do aluno\n(9 digitos de números inteiros): ")

                    if len(aux_matricula) > 9:
                        print("\n!ERRO!, insira SOMENTE 9 DIGITOS DE NÚMEROS INTEIROS!\n")
                        continue

                    if len(aux_matricula) < 9:
                        print("\n!ERRO!, insira 9 DIGITOS DE NÚMEROS INTEIROS!\n")
                        continue

                    if re.match(r'^[0-9]+$', aux_matricula): #novamente o re.match aparecendo aqui para somente permitir caractéres de 0 a 9
                        break

                    else:
                        print("\n!ERRO!, insira SOMENTE NÚMEROS!\n")

                except:
                    print("!ERRO!, insira SOMENTE NÚMEROS!")

            #insere o total de periodos integralizados
            while True:
                try:
                    aux_periodos = int(input("\nInsira o total de períodos cursados pelo aluno\n(número inteiro positivo:) "))
                    if 1 <= aux_periodos <= 11:
                        break
                    else:
                        print("\n!ERRO! insira um número, inteiro, maior ou igual a 1 e menor que 12!\n")
                except:
                    print("\n!ERRO! insira SOMENTE números inteiros positivos!\n")
            
            #insere as disciplinas e também excluí
            while True:
                try:
                    print("\n------MENU INSERE DISCIPLINA------\nDigite: 1 - Inserir disciplina e média final\nDigite: 2 - Excluir disciplina\nDigite: 3 - Sair\n")
                    x = int(input("Escolha uma opção: "))
                    
                    #inserir disciplinas
                    if x == 1:
                        for banco_disciplina in banco_disciplinas:
                            print("\n", banco_disciplina.codigo, " - ", banco_disciplina.nome)
                            
                        while True:
                            try:
                                aux_nome_codigo = input("\nInsira o código da disciplina:\n(Ou digite EXIT para sair): ")

                                if aux_nome_codigo in aux_codigo_media_disciplina:
                                    print("\n!ERRO! DISCIPLINA JÁ INSERIDA\n")

                                elif any(aux_nome_codigo == banco_disciplina.codigo for banco_disciplina in banco_disciplinas):
                                    while True:
                                        try:
                                            aux_media = float(input("Insira a média final da disciplina:\n(Entre 0 a 10): "))

                                            if 0.0 <= aux_media <= 10.0:
                                                aux_codigo_media_disciplina.update({aux_nome_codigo: aux_media})
                                                print("\n------Disciplina inserida com sucesso---")
                                                break

                                            else:
                                                print("\nERRO: Valor de média inválido, insira novamente\n")

                                        except:
                                            print("\nERRO: Valor de média inválido, insira novamente\n")

                                elif aux_nome_codigo == "EXIT":
                                    break

                                else:
                                    print("\n!ERRO! DISCIPLINA NÃO ENCONTRADA\n")

                            except:
                                pass
                
                    #excluir disciplinas
                    elif x == 2:
                        while True:
                            print(aux_codigo_media_disciplina)
                            codigo = input("\nInsira o código da disciplina que deseja excluir:\n(Ou digite EXIT para sair): ")
                            
                            if codigo == "EXIT":
                                break

                            else:
                                if codigo in aux_codigo_media_disciplina:
                                    del aux_codigo_media_disciplina[codigo]
                                    print("\n------Disciplina excluida com sucesso---")

                                else:
                                    print("\n!ERRO! DISCIPLINA NÃO ENCONTRADA\n")

                    #sai do programa
                    elif x == 3:
                        break

                    else:
                        print("\n!ERRO! OPÇÃO INVÁLIDA")

                except:
                    print("\n!ERRO! OPÇÃO INVÁLIDA")

            #instancia o aluno e adiciona na lista_alunos
            lista_alunos.append(Aluno(nome = aux_nome, matricula = aux_matricula, periodos_integralizados = aux_periodos, disciplinas_concluidas = aux_codigo_media_disciplina))
            print("\n")
            #atualiza QT_alunos para o próximo aluno, se existir
            QT_alunos_aux_controle += 1

    print(lista_alunos,"\n")



    nome_e1 = "COMPUTAÇÃO VISUAL"
    
    
    nome_e2 = "SISTEMAS INTELIGENTES"
   
    
    nome_e3 = "SISTEMAS DE COMPUTAÇÃO"
    
    
    nome_e4 = "SISTEMAS DE INFORMAÇÃO"
   
    
    escolha_enfase = {nome_e1 : ["COMP404", "COMP390", "COMP393", "COMP396", "COMP400"], nome_e2 : ["COMP404", "COMP390", "COMP393", "COMP397", "COMP401"], nome_e3 : ["COMP404", "COMP391", "COMP393", "COMP398", "COMP402"], nome_e4 : ["COMP389", "COMP392", "COMP395", "COMP399", "COMP403"]}
    



    controle_while = 0
    
    while(controle_while < QT_alunos):
        Escolhas = {1 : 0, 2 : 0, 3 : 0, 4 : 0}
        calculaNome1(lista_alunos[controle_while].nome, Escolhas)
        print("\n\n LINHA 699\n\n", Escolhas)
        
        lista_enfase = []
        
        for disciplina in banco_disciplinas:
            if disciplina.tipo == "Ênfase" and disciplina.codigo in lista_alunos[controle_while].disciplinas_concluidas:
                lista_enfase.append(disciplina.codigo)
        
        lista_disciplinas_a_fazer = []
        print("\n\nlinha 708\n\n:", lista_disciplinas_a_fazer)
        
        for disciplina in banco_disciplinas:
            if disciplina.tipo == 'Obrigatória' or disciplina.tipo == "Eletiva":
                lista_disciplinas_a_fazer.append(disciplina.codigo)
                    
        for item in lista_enfase:
            for codigo, nota in lista_alunos[controle_while].disciplinas_concluidas.items():
                if codigo == item and nota < 7.0:
                    lista_disciplinas_a_fazer.append(item)
        
        print("\n\nlinha 722\n\n:", lista_disciplinas_a_fazer)
        
        for codigo, nota in lista_alunos[controle_while].disciplinas_concluidas.items():
            if nota >= 7.0 and codigo in lista_disciplinas_a_fazer:
                lista_disciplinas_a_fazer.remove(codigo)
        
        list_tipoEnfase_feito = []

        for codigo, nota in lista_alunos[controle_while].disciplinas_concluidas.items():
            if nota >= 7.0:
                for disciplina in banco_disciplinas:
                    if disciplina.tipo == 'Ênfase' and disciplina.codigo == codigo:
                        list_tipoEnfase_feito.append(codigo)
                
        print("linha 730:", lista_disciplinas_a_fazer)
        
        
        inserePeriodo(Escolhas, escolha_enfase, lista_disciplinas_a_fazer, lista_alunos[controle_while].periodos_integralizados, lista_enfase, lista_alunos, list_tipoEnfase_feito)
        
        print("\nRestou disciplinas???:", lista_disciplinas_a_fazer)
        
        controle_while += 1


'''em python, a função main não é obrigatória, mas com o if __name__ == "__main__" vai verificar se o scrit está sendo executado diretamente'''
if __name__ == "__main__":
    main()

